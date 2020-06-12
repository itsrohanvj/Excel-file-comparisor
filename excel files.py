import xlrd
import openpyxl

import xlwt 
from xlwt import Workbook 

wb1 = openpyxl.load_workbook('file1.xlsx')
wb2 = openpyxl.load_workbook('file2.xlsx')
wb = Workbook()
sheet3 = wb.add_sheet('Sheet 1') 
sheet = wb1.get_sheet_by_name('Sheet1')
sheet2=wb2.get_sheet_by_name('Sheet1')
a=sheet.max_row
print(a)
b=sheet2.max_row
print(b)
sheet=wb1.active
sheet2=wb2.active
count=1
l1=[]
l2=[]
try:
    for k  in range(0,a):
    #here  column is fixed, row is dynamic.
        l1.append(sheet.cell(row=k+1,column=1).value)
    for k in range(0,b):
        l2.append(sheet2.cell(row=k+1,column=1).value)
    print (l1)
    print(l2)
    for i in range(0,min(a,b)):
        #compares every elements of sheets 1 with all the elements in sheet 2.
        if l1[i] in str(l2):
            continue
        else:
            print ("absent",l1[i]) #if the value is not present in sheet 2, it prints absent and appends that value to another excel sheet called 'final'.
            sheet3.write(count, 0, sheet.cell(row=i,column=1).value)
            wb.save('final.xls')
            count=count+1
            
except TypeError:
    
    print("Its fine")

